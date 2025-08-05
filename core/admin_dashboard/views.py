from functools import wraps
from openpyxl import Workbook
from django.db.models import Q
from orders.models import Order
from django.utils import timezone
from news.models.news import News
from blogs.models.blog import Blog
from django.db.models import Count
from rest_framework import generics
from django.shortcuts import render
from comments.models import Comment
from django.http import HttpResponse
from django.utils.timezone import now
from django.shortcuts import redirect
from logs.models import AdminActionLog
from django.core.mail import send_mail
from rest_framework.views import APIView
from payments.models import FailedPayment
from products.models.product import Product
from rest_framework.response import Response
from django.views.generic import TemplateView
from django.utils.dateparse import parse_datetime
from django.utils.crypto import get_random_string
from rest_framework.permissions import IsAuthenticated
from rest_framework.exceptions import PermissionDenied
from accounts.serializers.seller import SellerSerializer
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.decorators import user_passes_test, login_required
from accounts.serializers.admin import AdminPermissionSerializer
from accounts.models import User, AdminPermission, WriterPermission, AdminInviteToken
from accounts.decorators import admin_required, writer_required, seller_required, superadmin_required


# @login_required(login_url='/accounts/admin-login/')
# def superadmin_dashboard_view(request):
#     print("User:", request.user)
#     print("Is authenticated:", request.user.is_authenticated)
#     print("Is superuser:", request.user.is_superuser)
#     if not request.user.is_superuser:
#         return redirect('/admin-dashboard/')
#     return render(request, 'superadmin_dashboard/dashboard.html')

# @user_passes_test(lambda u: u.is_authenticated and u.is_superuser, login_url='/accounts/admin-login/')
# def superadmin_dashboard_view(request):
#     return render(request, 'superadmin_dashboard/dashboard.html')

@superadmin_required
def superadmin_dashboard_view(request):
    print("superadmin_dashboard_view")
    return render(request, 'superadmin_dashboard/dashboard.html')



@admin_required
def admin_dashboard_view(request):
    print("User:", request.user)
    print("Is authenticated:", request.user.is_authenticated)
    print("Is superuser:", request.user.is_superuser)
    if not request.user.is_superuser:
        return redirect('/superadmin-dashboard/')
    return render(request, 'admin_dashboard/dashboard.html')


@writer_required
def writer_dashboard_view(request):
    return render(request, 'writer_dashboard/dashboard.html')


@seller_required
def seller_dashboard_view(request):
    return render(request, 'seller_dashboard/dashboard.html')


class AdminSiteSummaryPage(LoginRequiredMixin, TemplateView):
    template_name = "admin_dashboard/summary.html"


class AdminSellersPage(LoginRequiredMixin, TemplateView):
    template_name = "admin_dashboard/sellers.html"


class AdminWritersPage(LoginRequiredMixin, TemplateView):
    template_name = "admin_dashboard/writers.html"


class AdminFlaggedCommentsPage(LoginRequiredMixin, TemplateView):
    template_name = "admin_dashboard/flagged_comments.html"


class AdminLowStockAlertsPage(LoginRequiredMixin, TemplateView):
    template_name = "admin_dashboard/low_stock.html"


class AdminFailedPaymentsPage(LoginRequiredMixin, TemplateView):
    template_name = "admin_dashboard/failed_payments.html"


class AdminCreateSellerPage(LoginRequiredMixin, TemplateView):
    template_name = "admin_dashboard/seller_create.html"


class AdminUpdateSellerPage(LoginRequiredMixin, TemplateView):
    template_name = "admin_dashboard/seller_update.html"


class AdminCreateAdminPage(LoginRequiredMixin, TemplateView):
    template_name = "admin_dashboard/admin_create.html"


class AdminLogsPage(LoginRequiredMixin, TemplateView):
    template_name = "admin_dashboard/admin_logs.html"


class AdminManageAdminsPage(LoginRequiredMixin, TemplateView):
    template_name = "admin_dashboard/manage_admins.html"


class AdminImpersonationPage(LoginRequiredMixin, TemplateView):
    template_name = "admin_dashboard/impersonation.html"


class AdminGrowthAnalyticsPage(LoginRequiredMixin, TemplateView):
    template_name = "admin_dashboard/growth_analytics.html"


class SiteSummaryView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if not request.user.adminpermission.is_superadmin and not request.user.adminpermission:
            return Response(status=403)
        return Response({
            "total_users": User.objects.count(),
            "total_sellers": User.objects.filter(role="seller").count(),
            "total_writers": User.objects.filter(writer_permission__isnull=False).count(),
            "total_products": Product.objects.count(),
            "total_orders": Order.objects.count(),
            "total_blogs": Blog.objects.count(),
            "total_news": News.objects.count(),
            "flagged_comments": Comment.objects.filter(status='flagged').count(),
        })


class SellersListView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if not hasattr(request.user, 'adminpermission'):
            return Response({"error": "Permission denied."}, status=403)

        sellers = User.objects.filter(role="seller")
        data = []
        for seller in sellers:
            product_count = Product.objects.filter(owner=seller).count()
            total_orders = Order.objects.filter(product__owner=seller).count()
            data.append({
                "id": seller.id,
                "name": f"{seller.first_name} {seller.last_name}",
                "phone": seller.phone_number,
                "email": seller.email,
                "product_count": product_count,
                "order_count": total_orders,
            })
        return Response(data)


class WritersListView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if not hasattr(request.user, 'adminpermission'):
            return Response({"error": "Permission denied."}, status=403)

        writers = User.objects.filter(writer_permission__isnull=False)
        data = []
        for writer in writers:
            perm = getattr(writer, 'writer_permission', None)
            blog_count = Blog.objects.filter(writer=writer).count()
            news_count = News.objects.filter(writer=writer).count()
            data.append({
                "id": writer.id,
                "name": f"{writer.first_name} {writer.last_name}",
                "phone": writer.phone_number,
                "email": writer.email,
                "can_write_blogs": perm.can_write_blogs if perm else False,
                "can_write_news": perm.can_write_news if perm else False,
                "blog_count": blog_count,
                "news_count": news_count,
            })
        return Response(data)


class FlaggedContentView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if not hasattr(request.user, 'adminpermission'):
            return Response({"error": "Permission denied."}, status=403)

        flagged = Comment.objects.filter(status='flagged').select_related('user')
        data = [
            {
                "id": c.id,
                "user": c.user.username,
                "content": c.content,
                "created": c.created_at,
                "status": c.status
            } for c in flagged
        ]
        return Response(data)

    def patch(self, request):
        if not hasattr(request.user, 'adminpermission'):
            return Response({"error": "Permission denied."}, status=403)

        cid = request.data.get("id")
        new_status = request.data.get("status")  # e.g. 'visible', 'hidden'
        try:
            comment = Comment.objects.get(id=cid)
            comment.status = new_status
            comment.save()
            return Response({"status": "updated"})
        except Comment.DoesNotExist:
            return Response({"error": "Comment not found."}, status=404)


class LowStockAlertsView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if not hasattr(request.user, 'adminpermission'):
            return Response({"error": "Permission denied."}, status=403)

        low_stock = Product.objects.filter(stock__lte=5)
        data = [
            {
                "id": p.id,
                "name": p.name,
                "stock": p.stock,
                "owner": p.owner.username
            } for p in low_stock
        ]
        return Response(data)


class FailedPaymentsView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if not hasattr(request.user, 'adminpermission'):
            return Response({"error": "Permission denied."}, status=403)

        failures = FailedPayment.objects.all().order_by('-created_at')[:50]
        data = [
            {
                "id": f.id,
                "user": f.user.username,
                "amount": f.amount,
                "reason": f.reason,
                "time": f.created_at
            } for f in failures
        ]
        return Response(data)


# Seller CRUD Views (Admins only)
# Inject logging into seller/writer/admin creation
class SellerCreateView(generics.CreateAPIView):
    queryset = User.objects.all()
    serializer_class = SellerSerializer
    permission_classes = [IsAuthenticated]

    def perform_create(self, serializer):
        instance = serializer.save(role="seller")
        AdminActionLogger.log(self.request.user, "Created seller", f"Seller ID: {instance.id} - {instance.email}")


class SellerUpdateDeleteView(generics.RetrieveUpdateDestroyAPIView):
    queryset = User.objects.filter(role="seller")
    serializer_class = SellerSerializer
    permission_classes = [IsAuthenticated]
    lookup_field = 'id'


# Writer Permission Management
class WriterPermissionUpdateView(APIView):
    permission_classes = [IsAuthenticated]

    def patch(self, request, writer_id):
        try:
            writer = User.objects.get(id=writer_id)
            perm, created = WriterPermission.objects.get_or_create(user=writer)
            perm.can_write_blogs = request.data.get("can_write_blogs", perm.can_write_blogs)
            perm.can_write_news = request.data.get("can_write_news", perm.can_write_news)
            perm.save()
            AdminActionLogger.log(request.user, "Updated writer permissions", f"Writer ID: {writer.id}")
            return Response({"status": "updated"})
        except User.DoesNotExist:
            return Response({"error": "Writer not found."}, status=404)


# Admin Management (Superadmin only)
class AdminCreateView(generics.CreateAPIView):
    serializer_class = AdminPermissionSerializer
    permission_classes = [IsAuthenticated]

    def perform_create(self, serializer):
        if not getattr(self.request.user, 'adminpermission',
                       None) or not self.request.user.adminpermission.is_superadmin:
            raise PermissionDenied("Only superadmins can add admins")
        instance = serializer.save()
        AdminActionLogger.log(self.request.user, "Created admin", f"Admin User ID: {instance.user.id}")


# CSV/Excel Export for Admin Reports
class ExportSellersExcelView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if not hasattr(request.user, 'adminpermission'):
            return Response({"error": "Permission denied."}, status=403)

        sellers = User.objects.filter(role="seller")
        wb = Workbook()
        ws = wb.active
        ws.append(["ID", "Name", "Phone", "Email"])
        for s in sellers:
            ws.append([s.id, f"{s.first_name} {s.last_name}", s.phone_number, s.email])
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=sellers.xlsx'
        wb.save(response)
        return response


class ExportWritersExcelView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if not hasattr(request.user, 'adminpermission'):
            return Response({"error": "Permission denied."}, status=403)

        writers = User.objects.filter(writer_permission__isnull=False)
        wb = Workbook()
        ws = wb.active
        ws.append(["ID", "Name", "Phone", "Email", "Can Write Blogs", "Can Write News"])
        for w in writers:
            perm = getattr(w, 'writer_permission', None)
            ws.append([
                w.id, f"{w.first_name} {w.last_name}", w.phone_number, w.email,
                perm.can_write_blogs if perm else False,
                perm.can_write_news if perm else False
            ])
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=writers.xlsx'
        wb.save(response)
        return response


class ExportFlaggedCommentsExcelView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if not hasattr(request.user, 'adminpermission'):
            return Response({"error": "Permission denied."}, status=403)

        flagged = Comment.objects.filter(status='flagged')
        wb = Workbook()
        ws = wb.active
        ws.append(["ID", "User", "Content", "Status", "Created"])
        for c in flagged:
            ws.append([c.id, c.user.username, c.content, c.status, str(c.created_at)])
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=flagged_comments.xlsx'
        wb.save(response)
        return response


# Filter/search users in admin dashboard
class FilteredUserListView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if not hasattr(request.user, 'adminpermission'):
            return Response({"error": "Permission denied."}, status=403)

        query = request.query_params.get("q", "")
        role = request.query_params.get("role")  # e.g. 'seller', 'writer', 'admin'

        users = User.objects.all()

        if query:
            users = users.filter(
                Q(first_name__icontains=query) |
                Q(last_name__icontains=query) |
                Q(email__icontains=query) |
                Q(phone_number__icontains=query)
            )

        if role == "seller":
            users = users.filter(role="seller")
        elif role == "writer":
            users = users.filter(writer_permission__isnull=False)
        elif role == "admin":
            users = users.filter(adminpermission__isnull=False)

        data = [
            {
                "id": u.id,
                "name": f"{u.first_name} {u.last_name}",
                "email": u.email,
                "phone": u.phone_number,
            } for u in users
        ]
        return Response(data)


# Admin Action Log - create function
class AdminActionLogger:
    @staticmethod
    def log(user, action, details=""):
        if hasattr(user, 'adminpermission'):
            AdminActionLog.objects.create(
                admin=user,
                action=action,
                details=details
            )


# Filter logs by admin username (GET param: ?admin=username)
class AdminLogsView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if not hasattr(request.user, 'adminpermission'):
            return Response({"error": "Permission denied."}, status=403)

        start = request.query_params.get("start")
        end = request.query_params.get("end")
        admin_name = request.query_params.get("admin")
        page = int(request.query_params.get("page", 1))
        page_size = int(request.query_params.get("page_size", 25))

        logs = AdminActionLog.objects.select_related('admin').order_by('-timestamp')

        if start:
            logs = logs.filter(timestamp__gte=parse_datetime(start))
        if end:
            logs = logs.filter(timestamp__lte=parse_datetime(end))
        if admin_name:
            logs = logs.filter(admin__username__icontains=admin_name)

        total = logs.count()
        start_idx = (page - 1) * page_size
        end_idx = start_idx + page_size
        logs = logs[start_idx:end_idx]

        data = [
            {
                "admin": log.admin.username,
                "action": log.action,
                "details": log.details,
                "time": log.timestamp
            } for log in logs
        ]
        return Response({"total": total, "results": data})


# Optional: integrate into frontend logger (example usage)
# In settings.py, use a custom DRF logger or log admin errors
# or forward logs to Sentry/Datadog via signal integration

class ExportFilteredLogsExcelView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if not hasattr(request.user, 'adminpermission'):
            return Response({"error": "Permission denied."}, status=403)

        start = request.query_params.get("start")
        end = request.query_params.get("end")
        admin_name = request.query_params.get("admin")

        logs = AdminActionLog.objects.select_related('admin').order_by('-timestamp')

        if start:
            logs = logs.filter(timestamp__gte=parse_datetime(start))
        if end:
            logs = logs.filter(timestamp__lte=parse_datetime(end))
        if admin_name:
            logs = logs.filter(admin__username__icontains=admin_name)

        wb = Workbook()
        ws = wb.active
        ws.append(["Admin", "Action", "Details", "Timestamp"])
        for log in logs:
            ws.append([log.admin.username, log.action, log.details, log.timestamp.strftime('%Y-%m-%d %H:%M:%S')])

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=filtered_admin_logs.xlsx'
        wb.save(response)
        return response


# Admin dashboard activity chart (e.g. frontend can use this for plotting)
class AdminActivitySummaryView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if not hasattr(request.user, 'adminpermission'):
            return Response({"error": "Permission denied."}, status=403)

        summary = (
            AdminActionLog.objects
            .extra(select={'date': "DATE(timestamp)"})
            .values('date')
            .annotate(count=Count('id'))
            .order_by('date')
        )
        return Response(summary)


# Sentry-style integration hook (pseudo-logger)
def log_admin_event_to_external_service(user, action, details):
    import requests
    if hasattr(user, 'adminpermission'):
        payload = {
            "username": user.username,
            "action": action,
            "details": details,
        }
        # Example: send to Slack webhook, Sentry, or internal logging microservice
        try:
            requests.post("https://hooks.slack.com/services/your/webhook/url", json=payload, timeout=3)
        except requests.exceptions.RequestException:
            pass


# Superadmin-only view for managing admins
class AdminManagementView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if not hasattr(request.user, 'adminpermission') or not request.user.adminpermission.is_superadmin:
            return Response({"error": "Permission denied."}, status=403)

        admins = AdminPermission.objects.select_related('user').all()
        data = [
            {
                "id": admin.user.id,
                "name": f"{admin.user.first_name} {admin.user.last_name}",
                "email": admin.user.email,
                "is_superadmin": admin.is_superadmin
            }
            for admin in admins
        ]
        return Response(data)

    def post(self, request):
        if not hasattr(request.user, 'adminpermission') or not request.user.adminpermission.is_superadmin:
            return Response({"error": "Permission denied."}, status=403)

        serializer = AdminPermissionSerializer(data=request.data)
        if serializer.is_valid():
            instance = serializer.save()
            AdminActionLogger.log(request.user, "Created admin", f"Admin ID: {instance.user.id}")
            return Response(serializer.data)
        return Response(serializer.errors, status=400)

    def patch(self, request):
        if not hasattr(request.user, 'adminpermission') or not request.user.adminpermission.is_superadmin:
            return Response({"error": "Permission denied."}, status=403)

        user_id = request.data.get('user_id')
        is_superadmin = request.data.get('is_superadmin')
        try:
            admin_perm = AdminPermission.objects.get(user__id=user_id)
            admin_perm.is_superadmin = is_superadmin
            admin_perm.save()
            AdminActionLogger.log(request.user, "Updated admin permission",
                                  f"Admin ID: {user_id}, Superadmin: {is_superadmin}")
            return Response({"status": "updated"})
        except AdminPermission.DoesNotExist:
            return Response({"error": "Admin not found."}, status=404)

    def delete(self, request):
        if not hasattr(request.user, 'adminpermission') or not request.user.adminpermission.is_superadmin:
            return Response({"error": "Permission denied."}, status=403)

        user_id = request.data.get('user_id')
        try:
            admin_perm = AdminPermission.objects.get(user__id=user_id)
            admin_perm.delete()
            AdminActionLogger.log(request.user, "Deleted admin", f"Admin ID: {user_id}")
            return Response({"status": "deleted"})
        except AdminPermission.DoesNotExist:
            return Response({"error": "Admin not found."}, status=404)


# Invite admin via email token (optional feature)
class InviteAdminView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        if not hasattr(request.user, 'adminpermission') or not request.user.adminpermission.is_superadmin:
            return Response({"error": "Permission denied."}, status=403)

        email = request.data.get("email")
        token = get_random_string(length=32)
        expiration = now() + timezone.timedelta(days=2)  # token valid for 48 hours

        AdminInviteToken.objects.create(email=email, token=token, expires_at=expiration)

        invite_link = f"https://your-site.com/admin-invite?token={token}"

        send_mail(
            subject="You’ve been invited as an Admin",
            message=f"Hello, click below to activate your admin role:\n{invite_link}",
            from_email="no-reply@your-site.com",
            recipient_list=[email],
            fail_silently=True
        )

        AdminActionLogger.log(request.user, "Sent admin invite", f"Invited Email: {email}")
        return Response({"status": "invite sent", "token": token})


# Admin visual analytics: count new admins over time
class AdminGrowthSummaryView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if not hasattr(request.user, 'adminpermission') or not request.user.adminpermission.is_superadmin:
            return Response({"error": "Permission denied."}, status=403)

        from django.db.models.functions import TruncDate
        from django.db.models import Count

        data = (
            AdminPermission.objects
            .annotate(date=TruncDate('created_at'))
            .values('date')
            .annotate(count=Count('id'))
            .order_by('date')
        )
        return Response(data)


# Optional feature: impersonate other admin (debugging only)
class ImpersonateAdminView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        if not hasattr(request.user, 'adminpermission') or not request.user.adminpermission.is_superadmin:
            return Response({"error": "Permission denied."}, status=403)

        user_id = request.data.get("user_id")
        try:
            target = User.objects.get(id=user_id)
            if not hasattr(target, 'adminpermission'):
                return Response({"error": "Target is not an admin"}, status=400)

            # Simulate impersonation (in real cases, generate token or session override)
            AdminActionLogger.log(request.user, "Impersonated admin", f"Now acting as: {target.id}")
            return Response({"message": f"Impersonating admin {target.email}"})

        except User.DoesNotExist:
            return Response({"error": "Admin not found."}, status=404)


# Activate admin via token
class ActivateAdminTokenView(APIView):
    def post(self, request):
        token = request.data.get("token")
        try:
            invite = AdminInviteToken.objects.get(token=token, expires_at__gte=now())
            user = User.objects.get(email=invite.email)
            AdminPermission.objects.get_or_create(user=user)
            invite.delete()
            AdminActionLogger.log(user, "Activated admin role via token")
            return Response({"status": "admin activated"})
        except (AdminInviteToken.DoesNotExist, User.DoesNotExist):
            return Response({"error": "Invalid or expired token."}, status=400)


# Suspend or expire admin access
class AdminSuspendView(APIView):
    permission_classes = [IsAuthenticated]

    def patch(self, request):
        if not hasattr(request.user, 'adminpermission') or not request.user.adminpermission.is_superadmin:
            return Response({"error": "Permission denied."}, status=403)

        user_id = request.data.get("user_id")
        suspend = request.data.get("suspend")
        expires_at = request.data.get("expires_at")

        try:
            admin = AdminPermission.objects.get(user__id=user_id)
            admin.is_suspended = suspend
            if expires_at:
                admin.expires_at = expires_at
            admin.save()
            AdminActionLogger.log(request.user, "Updated admin status",
                                  f"User ID: {user_id}, Suspended: {suspend}, Expires At: {expires_at}")
            return Response({"status": "admin updated"})
        except AdminPermission.DoesNotExist:
            return Response({"error": "Admin not found."}, status=404)


# Export admin audit logs to Excel
class ExportAdminAuditLogView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if not request.user.adminpermission.is_superadmin:
            return Response({"error": "Permission denied."}, status=403)

        logs = AdminActionLog.objects.select_related('admin').order_by('-timestamp')[:1000]

        wb = Workbook()
        ws = wb.active
        ws.append(["Admin", "Action", "Details", "Timestamp"])
        for log in logs:
            ws.append([log.admin.username, log.action, log.details, log.timestamp.strftime('%Y-%m-%d %H:%M:%S')])

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=admin_audit_log.xlsx'
        wb.save(response)
        return response
