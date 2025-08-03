from celery import shared_task
from django.core.mail import EmailMessage
from django.utils.timezone import now
from datetime import timedelta
import openpyxl
from openpyxl.utils import get_column_letter
from io import BytesIO

from accounts.models import User
from seller_dashboard.views import SellerDashboardSummaryView

@shared_task(name="send_weekly_report_task")
def send_weekly_report_task(user_id):
    try:
        user = User.objects.get(pk=user_id)
        start_date = now().date() - timedelta(days=7)
        end_date = now().date()

        summary = SellerDashboardSummaryView()
        summary_data = summary.get_summary_data(user, start_date, end_date)

        # Create Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Dashboard Summary"

        headers = list(summary_data.keys())
        values = list(summary_data.values())

        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws[f"{col_letter}1"] = header
            ws[f"{col_letter}2"] = values[col_num - 1]

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Send Email
        email = EmailMessage(
            subject="📈 Your Weekly Seller Dashboard",
            body="""
            <p>Hello,</p>
            <p>Your weekly dashboard summary is attached.</p>
            <p>Thanks,<br>The Team</p>
            """,
            to=[user.email],
        )
        email.content_subtype = "html"
        email.attach("dashboard_summary.xlsx", buffer.read(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        email.send()
    except Exception as e:
        print(f"❌ Error sending report to user {user_id}: {str(e)}")
