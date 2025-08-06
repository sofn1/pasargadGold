import csv
import json
import openpyxl
from io import BytesIO
from datetime import timedelta
from accounts.models import User
from django.core.cache import cache
from django.shortcuts import render
from django.db.models import Sum, F
from django.http import HttpResponse
from rest_framework.views import APIView
from django.core.mail import EmailMessage
from orders.models import Order, OrderItem
from products.models.product import Product
from rest_framework.response import Response
from openpyxl.utils import get_column_letter
from rest_framework.permissions import IsAuthenticated
from seller_dashboard.serializers import DateRangeFilterSerializer
from django_celery_beat.models import PeriodicTask, IntervalSchedule
from django.contrib.auth.decorators import login_required, user_passes_test


@login_required
@user_passes_test(lambda u: u.role == 'seller')
def seller_dashboard_home(request):
    return render(request, 'writer_dashboard/dashboard.html')


@login_required
def seller_summary_view(request):
    return render(request, 'seller_dashboard/summary.html')


@login_required
def revenue_time_series_view(request):
    return render(request, 'seller_dashboard/revenue_time_series.html')


@login_required
def order_time_series_view(request):
    return render(request, 'seller_dashboard/order_time_series.html')


@login_required
def profit_time_series_view(request):
    return render(request, 'seller_dashboard/profit_time_series.html')


@login_required
def top_products_view(request):
    return render(request, 'seller_dashboard/top_products.html')


@login_required
def low_stock_products_view(request):
    return render(request, 'seller_dashboard/low_stock_products.html')


@login_required
def order_funnel_view(request):
    return render(request, 'seller_dashboard/order_funnel.html')


@login_required
def geo_breakdown_view(request):
    return render(request, 'seller_dashboard/geo_breakdown.html')


@login_required
def device_breakdown_view(request):
    return render(request, 'seller_dashboard/device_breakdown.html')


@login_required
def unfulfilled_orders_view(request):
    return render(request, 'seller_dashboard/unfulfilled_orders.html')


@login_required
def system_health_view(request):
    return render(request, 'seller_dashboard/system_health.html')


@login_required
def sales_by_channel_view(request):
    return render(request, 'seller_dashboard/sales_by_channel.html')


class SellerDashboardSummaryView(APIView):
    permission_classes = [IsAuthenticated]

    def get_summary_data(self, user, start_date, end_date):
        cache_key = f"summary_{user.id}_{start_date}_{end_date}"
        cached = cache.get(cache_key)
        if cached:
            return cached

        orders = Order.objects.filter(user=user, status__in=['paid', 'sent', 'delivered'],
                                      created_at__range=(start_date, end_date))

        total_sales = orders.aggregate(Sum('total_price'))['total_price__sum'] or 0
        total_orders = orders.count()
        avg_order_value = total_sales / total_orders if total_orders else 0
        new_customers = User.objects.filter(date_joined__range=(start_date, end_date)).count()
        all_customers = User.objects.count()
        returning_customers = all_customers - new_customers

        cart_abandon_rate = 20.5
        net_profit = int(total_sales * 0.75)

        result = {
            "total_sales": total_sales,
            "net_profit": net_profit,
            "total_orders": total_orders,
            "average_order_value": avg_order_value,
            "cart_abandonment_rate": cart_abandon_rate,
            "new_customers": new_customers,
            "returning_customers": returning_customers,
        }
        cache.set(cache_key, result, 60 * 5)
        return result

    def get(self, request):
        user = request.user
        serializer = DateRangeFilterSerializer(data=request.query_params)
        serializer.is_valid(raise_exception=True)
        start_date, end_date = serializer.validated_data['start_date'], serializer.validated_data['end_date']
        result = self.get_summary_data(user, start_date, end_date)
        return Response(result)


class RevenueTimeSeriesView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        serializer = DateRangeFilterSerializer(data=request.query_params)
        serializer.is_valid(raise_exception=True)
        start_date, end_date = serializer.validated_data['start_date'], serializer.validated_data['end_date']

        delta = end_date - start_date
        data = []
        for i in range(delta.days + 1):
            day = start_date + timedelta(days=i)
            total = Order.objects.filter(
                user=user,
                status__in=['paid', 'sent', 'delivered'],
                created_at__date=day
            ).aggregate(Sum('total_price'))['total_price__sum'] or 0
            data.append({"day": str(day), "total": total})
        return Response(data)


class OrderTimeSeriesView(RevenueTimeSeriesView):
    def get(self, request):
        user = request.user
        serializer = DateRangeFilterSerializer(data=request.query_params)
        serializer.is_valid(raise_exception=True)
        start_date, end_date = serializer.validated_data['start_date'], serializer.validated_data['end_date']

        delta = end_date - start_date
        data = []
        for i in range(delta.days + 1):
            day = start_date + timedelta(days=i)
            count = Order.objects.filter(
                user=user,
                created_at__date=day,
                status__in=['paid', 'sent', 'delivered']
            ).count()
            data.append({"day": str(day), "orders": count})
        return Response(data)


class ProfitTimeSeriesView(RevenueTimeSeriesView):
    def get(self, request):
        user = request.user
        serializer = DateRangeFilterSerializer(data=request.query_params)
        serializer.is_valid(raise_exception=True)
        start_date, end_date = serializer.validated_data['start_date'], serializer.validated_data['end_date']

        delta = end_date - start_date
        data = []
        for i in range(delta.days + 1):
            day = start_date + timedelta(days=i)
            sales = Order.objects.filter(
                user=user,
                status__in=['paid', 'sent', 'delivered'],
                created_at__date=day
            ).aggregate(Sum('total_price'))['total_price__sum'] or 0
            profit = int(sales * 0.75)
            data.append({"day": str(day), "profit": profit})
        return Response(data)


class TopProductsView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        items = OrderItem.objects.filter(order__user=request.user)
        result = items.values('product__id', 'product__name').annotate(
            total_quantity=Sum('quantity'),
            total_revenue=Sum(F('quantity') * F('product__price'))
        ).order_by('-total_revenue')[:10]
        return Response(result)


class LowStockProductsView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        low_stock = Product.objects.filter(owner=request.user, stock__lte=10).values('id', 'name', 'stock')
        return Response(low_stock)


class OrderFunnelView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        funnel = {
            "visitors": 1500,
            "added_to_cart": 400,
            "checkout_started": 300,
            "payment_success": 220
        }
        return Response(funnel)


class GeoBreakdownView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        geo = [
            {"region": "Tehran", "orders": 50},
            {"region": "Mashhad", "orders": 30},
            {"region": "Isfahan", "orders": 20},
        ]
        return Response(geo)


class DeviceBreakdownView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        devices = [
            {"device": "mobile", "percentage": 70},
            {"device": "desktop", "percentage": 25},
            {"device": "tablet", "percentage": 5},
        ]
        return Response(devices)


class UnfulfilledOrdersView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        unfulfilled = Order.objects.filter(user=request.user, status='paid').values('id', 'total_price', 'created_at')
        return Response(unfulfilled)


class SystemHealthCheckView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        errors = Order.objects.filter(user=request.user, status='failed').count()
        return Response({"failed_orders": errors})


class SalesByChannelView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        product_id = request.query_params.get('product_id')
        if product_id:
            channels = [
                {"channel": "website", "sales": 2000000},
                {"channel": "instagram", "sales": 1000000},
                {"channel": "bazaar", "sales": 400000},
            ]
        else:
            channels = [
                {"channel": "website", "sales": 5500000},
                {"channel": "instagram", "sales": 2100000},
                {"channel": "bazaar", "sales": 800000},
            ]
        return Response(channels)


class ScheduleWeeklyReport(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        user = request.user
        schedule, _ = IntervalSchedule.objects.get_or_create(every=7, period=IntervalSchedule.DAYS)
        PeriodicTask.objects.create(
            interval=schedule,
            name=f"weekly_report_user_{user.id}",
            task="send_weekly_report_task",
            args=json.dumps([user.id])
        )
        return Response({"detail": "Weekly report scheduled."})


class EmailSummaryExcel(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        summary_view = SellerDashboardSummaryView()
        user = request.user
        serializer = DateRangeFilterSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        start_date, end_date = serializer.validated_data['start_date'], serializer.validated_data['end_date']
        response_data = summary_view.get_summary_data(user, start_date, end_date)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Dashboard Summary"

        headers = list(response_data.keys())
        values = list(response_data.values())

        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws[f"{col_letter}1"] = header
            ws[f"{col_letter}2"] = values[col_num - 1]

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        email = EmailMessage(
            subject="📊 Your Weekly Seller Dashboard Report",
            body="<p>Hello,</p><p>Your weekly performance summary is attached.</p><p>Best regards,<br/>Your Shop Team</p>",
            to=[user.email]
        )
        email.content_subtype = 'html'
        email.attach("dashboard_summary.xlsx", buffer.read(),
                     'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        email.send()

        return Response({"detail": "Report emailed successfully."})


class ExportSummaryExcel(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        summary_view = SellerDashboardSummaryView()
        user = request.user
        serializer = DateRangeFilterSerializer(data=request.query_params)
        serializer.is_valid(raise_exception=True)
        start_date, end_date = serializer.validated_data['start_date'], serializer.validated_data['end_date']
        response_data = summary_view.get_summary_data(user, start_date, end_date)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Dashboard Summary"

        headers = list(response_data.keys())
        values = list(response_data.values())

        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws[f"{col_letter}1"] = header
            ws[f"{col_letter}2"] = values[col_num - 1]

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(buffer,
                                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="dashboard_summary.xlsx"'
        return response


class ExportSummaryCSV(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        summary_view = SellerDashboardSummaryView()
        user = request.user
        serializer = DateRangeFilterSerializer(data=request.query_params)
        serializer.is_valid(raise_exception=True)
        start_date, end_date = serializer.validated_data['start_date'], serializer.validated_data['end_date']
        response_data = summary_view.get_summary_data(user, start_date, end_date)

        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="dashboard_summary.csv"'

        writer = csv.writer(response)
        writer.writerow(response_data.keys())
        writer.writerow(response_data.values())
        return response
