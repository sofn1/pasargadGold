import csv
from io import StringIO
from decimal import Decimal
from openpyxl import Workbook
from openpyxl import Workbook
from django.http import HttpResponse

def generate_excel_report(seller_data, filename="seller_report"):
    """
    Generate an Excel (.xlsx) file from seller data.

    Args:
        seller_data (dict or list of dicts): Data to export.
            Example:
                {
                    'summary': {
                        'Total Revenue': '120,000,000 تومان',
                        'Total Orders': 245,
                        ...
                    },
                    'top_products': [
                        {'name': 'Product A', 'sales': 50},
                        ...
                    ]
                }
        filename (str): Base name for the file (no extension)

    Returns:
        HttpResponse: Ready to return to user as attachment
    """
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "خلاصه فروش"

    # Add summary data
    if 'summary' in seller_data:
        ws_summary.append(["متغیر", "مقدار"])
        for key, value in seller_data['summary'].items():
            ws_summary.append([key, str(value)])

    # Add top products sheet
    if 'top_products' in seller_data:
        ws_products = wb.create_sheet(title="محصولات پرفروش")
        ws_products.append(["نام محصول", "تعداد فروش", "درآمد (تومان)"])
        for item in seller_data['top_products']:
            name = item.get('name', 'نامشخص')
            qty = item.get('quantity_sold', 0)
            revenue = item.get('revenue', Decimal('0'))
            ws_products.append([name, qty, str(revenue)])

    # Add low stock products
    if 'low_stock_products' in seller_data:
        ws_low = wb.create_sheet(title="کم‌موجودی")
        ws_low.append(["نام محصول", "موجودی فعلی", "حداقل موجودی"])
        for item in seller_data['low_stock_products']:
            ws_low.append([
                item.get('name', ''),
                item.get('current_stock', 0),
                item.get('min_stock', 0)
            ])

    # Save to HttpResponse
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
    wb.save(response)
    return response



def generate_csv_report(seller_data, filename="seller_report"):
    """
    Generate a CSV file from seller data.

    Args:
        seller_data (dict): Same structure as above
        filename (str): Base name for the file

    Returns:
        HttpResponse: CSV attachment
    """
    csv_buffer = StringIO()
    writer = csv.writer(csv_buffer)

    # Write summary
    writer.writerow(["بخش: خلاصه فروش"])
    writer.writerow(["متغیر", "مقدار"])
    if 'summary' in seller_data:
        for key, value in seller_data['summary'].items():
            writer.writerow([key, value])
    writer.writerow([])

    # Write top products
    writer.writerow(["بخش: محصولات پرفروش"])
    writer.writerow(["نام محصول", "تعداد فروش", "درآمد (تومان)"])
    if 'top_products' in seller_data:
        for item in seller_data['top_products']:
            writer.writerow([
                item.get('name', 'نامشخص'),
                item.get('quantity_sold', 0),
                item.get('revenue', 0)
            ])
    writer.writerow([])

    # Write low stock
    writer.writerow(["بخش: محصولات کم‌موجودی"])
    writer.writerow(["نام محصول", "موجودی فعلی", "حداقل موجودی"])
    if 'low_stock_products' in seller_data:
        for item in seller_data['low_stock_products']:
            writer.writerow([
                item.get('name', ''),
                item.get('current_stock', 0),
                item.get('min_stock', 0)
            ])

    # Return as HttpResponse
    response = HttpResponse(csv_buffer.getvalue().encode('utf-8'), content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
    return response


def generate_excel_http_response(data, filename="report.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    # Header
    ws.append(["Metric", "Value"])
    for key, value in data.items():
        ws.append([key, str(value)])

    # Response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response